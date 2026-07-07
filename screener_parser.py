"""
screener_parser.py
Generic reader for Screener.in "Export to Excel" workbooks.
Every screener.in export follows the same fixed layout on the 'Data Sheet'
tab (company meta at top, then PROFIT & LOSS / Quarters / BALANCE SHEET /
CASH FLOW / PRICE / DERIVED blocks, each with a 'Report Date' header row
followed by labelled line items). This parser locates each block by its
marker text so it works for ANY company export, not just TCS, regardless
of how many years of history are present or which optional line items
(e.g. Power & Fuel) are populated.
"""
import openpyxl
from datetime import datetime


def _row_values(ws, row, col_start, col_end):
    return [ws.cell(row=row, column=c).value for c in range(col_start, col_end + 1)]


def _find_label_row(ws, label, start_row=1, end_row=None, col=1):
    end_row = end_row or ws.max_row
    for r in range(start_row, end_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and str(v).strip().lower() == label.lower():
            return r
    return None


def _read_block(ws, header_row, labels, ncols):
    """Read a block of {label: [values...]} starting the row after a
    'Report Date' header row, for the given number of data columns
    (columns B onward)."""
    dates = _row_values(ws, header_row, 2, 1 + ncols)
    data = {}
    r = header_row + 1
    while r <= ws.max_row:
        label = ws.cell(row=r, column=1).value
        if label is None:
            break
        label = str(label).strip()
        data[label] = _row_values(ws, r, 2, 1 + ncols)
        r += 1
    return dates, data


def parse_screener_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Data Sheet" not in wb.sheetnames:
        raise ValueError("This does not look like a Screener.in export — no 'Data Sheet' tab found.")
    ws = wb["Data Sheet"]

    # ---- META ----
    company_name = ws["B1"].value
    face_value = ws["B7"].value or 1.0
    current_price = ws["B8"].value
    market_cap = ws["B9"].value

    # ---- PROFIT & LOSS (annual) ----
    pl_marker = _find_label_row(ws, "PROFIT & LOSS")
    pl_date_row = _find_label_row(ws, "Report Date", start_row=pl_marker, end_row=pl_marker + 3)
    pl_dates_raw, pl_raw = _read_block(ws, pl_date_row, None, ncols=10)
    # Younger companies (fewer than 10 listed years) get blank columns padded at
    # the START, not the end, so find the actual populated columns rather than
    # assuming they're the first N.
    valid_idx = [i for i, d in enumerate(pl_dates_raw) if isinstance(d, datetime)]
    if not valid_idx:
        raise ValueError("Could not find any valid report dates in the PROFIT & LOSS block.")
    n_years = len(valid_idx)
    start_idx = valid_idx[0]
    pl_dates = [pl_dates_raw[i] for i in valid_idx]
    pl = {k: [v[i] for i in valid_idx] for k, v in pl_raw.items()}

    # ---- BALANCE SHEET ----
    bs_marker = _find_label_row(ws, "BALANCE SHEET")
    bs_date_row = _find_label_row(ws, "Report Date", start_row=bs_marker, end_row=bs_marker + 3)
    _, bs_raw = _read_block(ws, bs_date_row, None, ncols=10)
    bs = {k: v[start_idx:start_idx + n_years] for k, v in bs_raw.items()}

    # ---- CASH FLOW ----
    cf_marker = _find_label_row(ws, "CASH FLOW:")
    cf_date_row = _find_label_row(ws, "Report Date", start_row=cf_marker, end_row=cf_marker + 3)
    _, cf_raw = _read_block(ws, cf_date_row, None, ncols=10)
    cf = {k: v[start_idx:start_idx + n_years] for k, v in cf_raw.items()}

    # ---- DERIVED (adjusted shares outstanding series) ----
    derived_marker = _find_label_row(ws, "DERIVED:")
    shares_row = None
    if derived_marker:
        shares_row = _find_label_row(ws, "Adjusted Equity Shares in Cr",
                                      start_row=derived_marker, end_row=derived_marker + 5)
    shares_series_raw = _row_values(ws, shares_row, 2, 11) if shares_row else [None] * 10
    shares_series = shares_series_raw[start_idx:start_idx + n_years] if shares_row else [None] * n_years
    shares_outstanding_latest = shares_series[-1] if shares_series and shares_series[-1] else market_cap / current_price

    def g(block, label):
        vals = block.get(label)
        if vals is None:
            return [0] * n_years
        return [v if isinstance(v, (int, float)) else 0 for v in vals]

    # Fiscal year labels e.g. FY2026 from the last date in the P&L block
    fy_end_years = [d.year if d.month >= 4 else d.year for d in pl_dates]  # Indian FY label = calendar year of March-end

    result = {
        "company_name": str(company_name).strip() if company_name else "Unknown Company",
        "face_value": face_value,
        "current_price": current_price,
        "market_cap": market_cap,
        "n_years": n_years,
        "fy_labels": [f"FY{y}" for y in fy_end_years],
        "shares_outstanding_series": [s if isinstance(s, (int, float)) else None for s in shares_series],
        "shares_outstanding_latest": shares_outstanding_latest,

        # P&L
        "sales": g(pl, "Sales"),
        "raw_material": g(pl, "Raw Material Cost"),
        "power_fuel": g(pl, "Power and Fuel"),
        "other_mfr_exp": g(pl, "Other Mfr. Exp"),
        "employee_cost": g(pl, "Employee Cost"),
        "selling_admin": g(pl, "Selling and admin"),
        "other_expenses": g(pl, "Other Expenses"),
        "other_income": g(pl, "Other Income"),
        "depreciation": g(pl, "Depreciation"),
        "interest": g(pl, "Interest"),
        "tax": g(pl, "Tax"),
        "dividend_amount": g(pl, "Dividend Amount"),

        # Balance sheet
        "equity_share_capital": g(bs, "Equity Share Capital"),
        "reserves": g(bs, "Reserves"),
        "borrowings": g(bs, "Borrowings"),
        "other_liabilities": g(bs, "Other Liabilities"),
        "total_liabilities": g(bs, "Total"),
        "net_block": g(bs, "Net Block"),
        "cwip": g(bs, "Capital Work in Progress"),
        "investments": g(bs, "Investments"),
        "total_assets": g(bs, "Total"),
        "receivables": g(bs, "Receivables"),
        "inventory": g(bs, "Inventory"),
        "cash_bank": g(bs, "Cash & Bank"),

        # Cash flow
        "cfo": g(cf, "Cash from Operating Activity"),
        "cfi": g(cf, "Cash from Investing Activity"),
        "cff": g(cf, "Cash from Financing Activity"),
    }

    # Derived line: Cost of Materials & Other Direct Costs = Raw Material + Power&Fuel + Other Mfr Exp
    result["cost_of_materials"] = [
        (result["raw_material"][i] or 0) + (result["power_fuel"][i] or 0) + (result["other_mfr_exp"][i] or 0)
        for i in range(n_years)
    ]
    # Screener's Balance-Sheet "Other Assets" line is a ROLLED-UP figure that
    # already includes Receivables + Inventory + Cash & Bank (those are shown
    # again below merely as working-capital memo items). Net it down to the
    # true residual "other assets" (unbilled revenue, prepaid, DTA, etc.) so
    # it doesn't double-count against the separate BS lines for those items.
    raw_other_assets = g(bs, "Other Assets")
    result["other_assets"] = [
        (raw_other_assets[i] or 0) - (result["receivables"][i] or 0)
        - (result["inventory"][i] or 0) - (result["cash_bank"][i] or 0)
        for i in range(n_years)
    ]
    return result


if __name__ == "__main__":
    import json, sys
    data = parse_screener_file(sys.argv[1] if len(sys.argv) > 1 else "tcs_screener.xlsx")
    print(json.dumps({k: v for k, v in data.items() if not isinstance(v, list) or len(v) < 15}, indent=2, default=str))
    print("Years:", data["fy_labels"])
    print("Sales:", data["sales"])