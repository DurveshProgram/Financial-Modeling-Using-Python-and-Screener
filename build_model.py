"""
build_model.py
Takes parsed Screener.in data + macro assumptions and populates the
professional three-statement model template (Cover / Assumptions / IS /
BS / CF / Ratios / DCF / Comps / Dashboard), producing a fully-linked,
company-specific Excel model.

Design principles:
- Historical actuals (columns B:K, i.e. the years actually reported by
  Screener.in) are hardcoded BLUE inputs, sourced straight from the
  uploaded screener file.
- All ratios, margins, DCF math, comps math etc. remain live FORMULAS
  (black/green per model conventions) -- untouched from the template,
  since that logic is company-agnostic.
- Forward projection assumptions (columns L:P) default to the AVERAGE of
  the last five available actual years (or however many are available, if
  fewer than five) for that line item, each still a plain, editable input
  cell -- NOT a python-computed hardcode -- so changing the historical
  data or the assumption cell itself recalculates the whole model inside
  Excel.
- Cost-of-capital inputs (risk-free rate, equity risk premium, beta) and
  the Comps-sheet peer set / sector classification are populated from
  live Yahoo Finance data via market_data.py, with the fetch date and
  source noted in a cell comment. If a live fetch fails (no network /
  yfinance not installed), a clearly labelled fallback value is used
  instead so the workbook still builds.
"""
import openpyxl
from openpyxl.styles import Font
from openpyxl.comments import Comment

BLUE = Font(color="0000FF")
COLS = list("BCDEFGHIJKLMNOP")  # B..K = 10 historical, L..P = 5 projected


def _set(ws, row, col_letter, value, blue=True, comment=None):
    cell = ws[f"{col_letter}{row}"]
    cell.value = value
    if blue:
        cell.font = Font(name=cell.font.name or "Calibri", color="0000FF")
    if comment:
        cell.comment = Comment(comment, "Model Engine")


def _fill_historical_row(ws, row, values, n_hist, source_note=None):
    for i, v in enumerate(values[-n_hist:]):
        col = COLS[i]
        _set(ws, row, col, v if v is not None else 0, blue=True,
             comment=source_note if i == 0 else None)


def _avg_last5_projection(ws, row, n_hist, n_proj=5):
    """Each projection column defaults to the AVERAGE of the last five
    available actual years for that row (or all available actuals if
    fewer than five years of history exist). Every projection column gets
    its own independent formula (not chained to the previous projected
    column), so overriding one forecast year doesn't silently change the
    others. IFERROR falls back to the last actual value if the average
    can't be computed (e.g. a row with no populated history at all)."""
    lookback = min(5, n_hist)
    first_col = COLS[n_hist - lookback]
    last_col = COLS[n_hist - 1]
    hist_range = f"{first_col}{row}:{last_col}{row}"
    proj_cols = COLS[n_hist:n_hist + n_proj]
    for c in proj_cols:
        ws[f"{c}{row}"] = f"=IFERROR(AVERAGE({hist_range}),{last_col}{row})"


def build_model(template_path, output_path, data, macro, market=None):
    """market: optional dict from generate_model.py / market_data.py with
    keys: ticker, sector_key, index_label, profile (Yahoo .info subset),
    peers (list of peer-multiple dicts). If None or incomplete, sensible
    generic fallbacks are used instead of hardcoding IT/TCS specifics."""
    market = market or {}
    wb = openpyxl.load_workbook(template_path)
    n = data["n_years"]
    fy = data["fy_labels"]
    proj_labels = [f"FY{int(fy[-1][2:]) + i}E" for i in range(1, 6)]
    all_labels = fy + proj_labels
    company = data["company_name"]

    profile = market.get("profile", {})
    index_label = market.get("index_label", "Sector / index not resolved — set manually")
    sector_display = (profile.get("sector") or profile.get("industry")
                       or (index_label.replace("Nifty ", "") if market.get("sector_key") else None)
                       or "Sector not classified")

    # ---------------- COVER ----------------
    cov = wb["Cover"]
    cov["C3"] = f"{company.upper()} — INTEGRATED FINANCIAL MODEL & INTRINSIC VALUATION"
    cov["E14"] = data["current_price"]
    cov["E15"] = data["shares_outstanding_latest"]
    cov["E16"] = round(data["current_price"] * data["shares_outstanding_latest"], 2)
    cov["C9"] = "Prepared: " + macro["prepared_date"]
    lo, hi = profile.get("fiftyTwoWeekLow"), profile.get("fiftyTwoWeekHigh")
    sources_used = profile.get("sources_used", [])
    if lo and hi:
        cov["E17"] = f"₹{lo:.2f}–₹{hi:.2f} / {sector_display} — {index_label}"
        cov["E17"].comment = Comment(
            f"LIVE: {' + '.join(sources_used) if sources_used else 'market data'} "
            f"({market.get('ticker') or market.get('nse_symbol') or 'n/a'}), 52-week range "
            f"and sector/industry classification, as of {macro['prepared_date']}.",
            "Model Engine")
    else:
        cov["E17"] = f"{sector_display} — {index_label}"
        cov["E17"].comment = Comment(
            "Sector classification and/or 52-week range could not be fetched live "
            "(NSE/TradingView/Yahoo/Google Finance all unavailable/blocked, or ticker not resolved) "
            "— verify and override manually.",
            "Model Engine")

    # ---------------- Year headers on every statement sheet ----------------
    for sheet_name, header_rows in [("Assumptions", [5]), ("IS", [5]), ("BS", [5]),
                                     ("CF", [5, 14]), ("Ratios", [5]), ("DCF", []),
                                     ]:
        ws = wb[sheet_name]
        ws["A2"] = f"{company} (Consolidated)"
        for hr in header_rows:
            for i, col in enumerate(COLS[:len(all_labels)]):
                ws[f"{col}{hr}"] = all_labels[i]

    for sheet_name in ["IS", "BS", "CF", "Ratios"]:
        wb[sheet_name]["A2"] = f"{company} (Consolidated)"
    wb["Dashboard"]["A1"] = f"{company.upper()} — VALUATION DASHBOARD"

    # ---------------- Comps: sector, company name, peer multiples ----------------
    comps = wb["Comps"]
    if market.get("sector_key") is None:
        comps["A2"] = "Sector not auto-classified — peer set requires manual entry (see row 7 comment)"
    else:
        comps["A2"] = f"Trailing multiples, {sector_display} ({index_label}) peer set"
    comps["A6"] = company
    comps["A10"] = f"Peer Average (excl. {company})"
    comps["A12"] = f"IMPLIED VALUATION — {company.upper()} AT PEER-AVERAGE MULTIPLES"
    comps["C13"] = f"{company} Metric ({fy[-1]}A, ₹ Cr)"

    peers = market.get("peers", [])
    peer_rows = [7, 8, 9]  # template has 3 peer slots

    # Always clear all 3 peer slots FIRST, regardless of how many live peers
    # were found. Previously, if fewer than 3 peers came back (e.g. because
    # the company excluded itself from its own 3-name sector list), the
    # template's original static placeholder data silently survived in the
    # unfilled row(s) and looked like real data for the new company.
    for r in peer_rows:
        for col in "ABCDEF":
            comps[f"{col}{r}"] = None
            comps[f"{col}{r}"].comment = None

    if peers:
        for row_idx, peer in zip(peer_rows, peers[:3]):
            r = row_idx
            comps[f"A{r}"] = f"{peer['name']} ({peer['ticker'].replace('.NS', '')})"
            if peer.get("live"):
                fetch_note = f"LIVE: {peer.get('source', 'market data')}, as of {macro['prepared_date']}."
            elif peer.get("source"):
                fetch_note = (f"PARTIAL — {peer['source']}, as of {macro['prepared_date']}. Some fields "
                               "could not be fetched from any source (blank = excluded from the peer "
                               "average, not treated as zero); verify manually before relying on the "
                               "peer-average valuation.")
            else:
                fetch_note = ("FALLBACK — live fetch failed for this peer on every source (TradingView, "
                               "Yahoo, Google Finance all unavailable/blocked); verify this peer's "
                               "multiples manually before relying on the peer-average valuation.")
            # None (not 0) for anything unfetched -- a blank cell is excluded
            # from AVERAGE(); a 0 would silently corrupt the peer average and
            # look like a real (very cheap) multiple.
            _set(comps, r, "B", peer.get("mkt_cap_cr"), blue=True, comment=fetch_note)
            _set(comps, r, "C", peer.get("pe"), blue=True)
            _set(comps, r, "D", peer.get("ev_ebitda"), blue=True)
            _set(comps, r, "E", peer.get("pb"), blue=True)
            _set(comps, r, "F", peer.get("roe"), blue=True)
    else:
        reason = ("this company's sector could not be classified (no name-keyword match and "
                   "no live NSE/Yahoo sector data)" if market.get("sector_key") is None else
                   "live peer data could not be fetched (NSE/TradingView/Yahoo/Google Finance all "
                   "unavailable or blocked)")
        for i, r in enumerate(peer_rows):
            comps[f"A{r}"] = f"[Enter Peer {i+1} name — {reason}]"
        comps["A7"].comment = Comment(
            f"Peer set not auto-populated: {reason}. Replace these 3 placeholder rows with "
            "2-4 actual listed peers for this company's real sector (name, market cap, "
            "trailing P/E, EV/EBITDA, P/B, ROE) before relying on the peer-multiple "
            "valuation on this sheet or the Dashboard. Ask Claude to look up current peer "
            "multiples for this company's sector, or fill in from your own data terminal.",
            "Model Engine")

    # ---------------- IS: historical actuals ----------------
    isws = wb["IS"]
    src = "Source: Screener.in export, uploaded by user"
    _fill_historical_row(isws, 6, data["sales"], n, src)
    _fill_historical_row(isws, 10, data["employee_cost"], n)
    _fill_historical_row(isws, 11, data["cost_of_materials"], n,
                          "= Raw Material + Power & Fuel + Other Mfg. Exp. (Screener.in)")
    _fill_historical_row(isws, 12, data["selling_admin"], n)
    _fill_historical_row(isws, 13, data["other_expenses"], n)
    _fill_historical_row(isws, 17, data["other_income"], n)
    _fill_historical_row(isws, 18, data["depreciation"], n)
    _fill_historical_row(isws, 20, data["interest"], n)
    _fill_historical_row(isws, 22, data["tax"], n)
    _fill_historical_row(isws, 26, data["dividend_amount"], n)
    _fill_historical_row(isws, 30, [s if s else data["shares_outstanding_latest"]
                                     for s in data["shares_outstanding_series"]], n)

    # ---------------- BS: historical actuals ----------------
    bsws = wb["BS"]
    _fill_historical_row(bsws, 6, data["net_block"], n, src)
    _fill_historical_row(bsws, 7, data["cwip"], n)
    _fill_historical_row(bsws, 8, data["investments"], n)
    _fill_historical_row(bsws, 9, data["receivables"], n)
    _fill_historical_row(bsws, 10, data["inventory"], n)
    _fill_historical_row(bsws, 11, data["cash_bank"], n)
    _fill_historical_row(bsws, 12, data["other_assets"], n)
    _fill_historical_row(bsws, 16, data["equity_share_capital"], n)
    _fill_historical_row(bsws, 17, data["reserves"], n)
    _fill_historical_row(bsws, 18, data["borrowings"], n)
    _fill_historical_row(bsws, 19, data["other_liabilities"], n)

    # ---------------- CF: historical actuals (reported) ----------------
    cfws = wb["CF"]
    _fill_historical_row(cfws, 6, data["cfo"], n, src)
    _fill_historical_row(cfws, 7, data["cfi"], n)
    _fill_historical_row(cfws, 8, data["cff"], n)

    # ---------------- Ratios: historical market price ----------------
    # NOTE: this row drives historical P/E, P/B, and EV/EBITDA in the Ratios
    # sheet. It was previously left untouched by the model engine, so every
    # non-TCS company silently inherited TCS's actual historical share price
    # here -- producing wrong historical valuation multiples for anyone else.
    rws = wb["Ratios"]
    _fill_historical_row(rws, 31, data["price_history"], n, src)

    # ---------------- Assumptions: WACC / macro block ----------------
    aws = wb["Assumptions"]
    aws["A1"] = f"{company} — Model Assumptions & Cost of Capital"
    _set(aws, 19, "C", macro["risk_free_rate"], comment=macro["risk_free_rate_source"])
    _set(aws, 20, "C", macro["equity_risk_premium"], comment=macro["erp_source"])
    _set(aws, 21, "C", macro["beta"], comment=macro["beta_source"])
    _set(aws, 23, "C", macro["pretax_cost_of_debt"], comment="Default: FY latest effective interest rate on borrowings — override with issuer's actual cost of debt/credit spread if known.")
    _set(aws, 32, "C", macro["terminal_growth"], comment=macro["terminal_growth_source"])
    _set(aws, 35, "C", data["current_price"], comment=src)
    _set(aws, 37, "C", data["shares_outstanding_latest"], comment=src)
    _set(aws, 38, "C", data["face_value"], comment=src)

    # ---------------- Assumptions: projection defaults = avg of last 5 yrs ----------------
    for row in range(6, 17):
        _avg_last5_projection(aws, row, n)
        for c in COLS[n:n + 5]:
            aws[f"{c}{row}"].font = Font(color="0000FF")

    return wb, all_labels